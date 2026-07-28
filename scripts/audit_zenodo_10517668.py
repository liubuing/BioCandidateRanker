from __future__ import annotations

import hashlib
import json
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import curve_fit


ROOT = Path(__file__).resolve().parents[1]
SOURCE = (
    ROOT
    / "artifacts"
    / "external"
    / "temporal-absolute-kinetics"
    / "zenodo-10517668"
)
RAW = SOURCE / "raw"
WORKBOOK = RAW / "Fig2_raw_data.xlsx"
TRAINING_FASTA = RAW / "training_seqs.fasta"
METADATA = RAW / "metadata.json"

EXPECTED_SHA256 = {
    "Fig2_raw_data.xlsx": "3a2dcad4e5d2dbe702e0a4328ee20195078084682eb61f50266cadf8dba19167",
    "training_seqs.fasta": "670ac87947ff51e70fb7ae606ad5ca0485e191c738f007e38ab1d5d507653db8",
    "metadata.json": "ede03439645c7908ae66b0b0f2492c5c581c7d1e39aebdf338edfb110dba9656",
}
ENZYMES = {
    "Pf-K": ("B", 4),
    "K1": ("E", 3),
    "K2": ("H", 3),
    "K3": ("K", 4),
    "K31": ("N", 3),
}
CONCENTRATIONS_UM = np.array([25, 50, 100, 200, 400, 700], dtype=float)
REPLICATE_START_ROWS = (22, 27, 32, 37, 42, 47)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_fasta(path: Path) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    header = ""
    sequence: list[str] = []
    for line in path.read_text(encoding="ascii").splitlines():
        if line.startswith(">"):
            if header:
                records.append((header, "".join(sequence)))
            header = line[1:]
            sequence = []
        else:
            sequence.append(line.strip())
    if header:
        records.append((header, "".join(sequence)))
    return records


def michaelis_menten(substrate: np.ndarray, kcat: float, km: float) -> np.ndarray:
    return kcat * substrate / (km + substrate)


def fit(substrate: np.ndarray, rates: np.ndarray, sigma: np.ndarray | None = None) -> dict[str, float]:
    parameters, _ = curve_fit(
        michaelis_menten,
        substrate,
        rates,
        p0=(float(max(rates)), 100.0),
        bounds=(0, np.inf),
        sigma=sigma,
        absolute_sigma=sigma is not None,
        maxfev=100_000,
    )
    return {"kcat_s-1": float(parameters[0]), "km_uM": float(parameters[1])}


def workbook_audit() -> tuple[dict[str, object], dict[str, list[list[float]]]]:
    formulas = load_workbook(WORKBOOK, data_only=False, read_only=False)["Fig. 2b"]
    values = load_workbook(WORKBOOK, data_only=True, read_only=False)["Fig. 2b"]
    results: dict[str, object] = {}
    raw_replicates: dict[str, list[list[float]]] = {}
    for enzyme, (column, replicate_count) in ENZYMES.items():
        replicates = [
            [float(values[f"{column}{row + offset}"].value) for offset in range(replicate_count)]
            for row in REPLICATE_START_ROWS
        ]
        raw_replicates[enzyme] = replicates
        arrays = [np.asarray(group) for group in replicates]
        means = np.asarray([group.mean() for group in arrays])
        standard_deviations = np.asarray([group.std(ddof=1) for group in arrays])
        pooled = np.concatenate(arrays)
        results[enzyme] = {
            "replicates_per_nonzero_concentration": replicate_count,
            "mean_formula_cells": [
                formulas[f"{column}{row - 1}"].value for row in REPLICATE_START_ROWS
            ],
            "diagnostic_fits_not_curated_labels": {
                "unweighted_concentration_means": fit(CONCENTRATIONS_UM, means),
                "unweighted_individual_replicates": fit(
                    np.repeat(CONCENTRATIONS_UM, replicate_count), pooled
                ),
                "inverse_sd_weighted_concentration_means": fit(
                    CONCENTRATIONS_UM, means, standard_deviations
                ),
            },
        }
    return results, raw_replicates


def main() -> None:
    observed_hashes = {name: sha256(RAW / name) for name in EXPECTED_SHA256}
    if observed_hashes != EXPECTED_SHA256:
        raise ValueError(f"Source hash mismatch: {observed_hashes}")

    metadata = json.loads(METADATA.read_text(encoding="utf-8"))
    license_id = metadata["metadata"]["license"]["id"]
    fasta_records = read_fasta(TRAINING_FASTA)
    headers = [header for header, _ in fasta_records]
    discovery_headers = [
        header for header in headers if header.split("|", 1)[0] in {"K1", "K2", "K3", "K31"}
    ]
    pf_records = [record for record in fasta_records if record[0].startswith("Pf-KYNU|")]
    if len(pf_records) != 1:
        raise ValueError(f"Expected one Pf-KYNU record, found {len(pf_records)}")

    homology_dir = SOURCE / "homology"
    homology_dir.mkdir(exist_ok=True)
    (homology_dir / "pf-k.fasta").write_text(
        f">Pf-K|WP_017531066.1\n{pf_records[0][1]}\n", encoding="ascii"
    )

    fit_audit, raw_replicates = workbook_audit()
    output = {
        "schema_version": 1,
        "source_id": "zenodo-10517668",
        "dataset_doi": "10.5281/zenodo.10517668",
        "article_doi": "10.1093/nar/gkae1245",
        "protocol_frozen_on": "2026-07-20",
        "status": "scientifically_blocked_no_curated_records",
        "source_integrity": {
            "sha256": observed_hashes,
            "dataset_license": license_id,
            "license_gate_passes": license_id == "cc-by-4.0",
            "article_published": "2025-01-07",
            "temporal_gate_passes": True,
        },
        "workbook_evidence": {
            "sheet": "Fig. 2b",
            "substrate": "L-kynurenine",
            "substrate_concentrations_uM": CONCENTRATIONS_UM.tolist(),
            "zero_concentration_summary_only": True,
            "raw_nonzero_replicates": raw_replicates,
            "author_method": "Michaelis-Menten fit using Origin; no weighting, replicate-pooling, zero-handling, optimizer, or uncertainty settings declared",
            "fit_sensitivity_audit": fit_audit,
            "fit_values_are_candidate_labels": False,
        },
        "sequence_evidence": {
            "training_fasta_records": len(fasta_records),
            "training_fasta_unique_sequences": len({sequence for _, sequence in fasta_records}),
            "discovery_label_headers": discovery_headers,
            "pf_k_mapping": {
                "workbook_label": "Pf-K",
                "fasta_header": pf_records[0][0],
                "article_accession": "WP_017531066.1",
                "sequence_sha256": hashlib.sha256(pf_records[0][1].encode("ascii")).hexdigest(),
                "mapping_basis": "article accession sequence is byte-identical to deposited Pf-KYNU sequence",
            },
            "unmapped_labels": ["K1", "K2", "K3", "K31"],
            "no_guessing_statement": "The training FASTA is the legacy 159-record regressor corpus and contains no K1, K2, K3, or K31 identifier; sequence similarity or rank order was not used to assign labels.",
        },
        "blocking_gates": [
            {
                "gate": "predeclared_reproducible_direct_kcat_fit",
                "affected_labels": list(ENZYMES),
                "reason_code": "fit_method_underdeclared",
                "reason": "Raw replicates permit multiple scientifically reasonable fits, but the source does not predeclare enough details to select one without post hoc analyst choice. Weighted and unweighted fits differ materially for some enzymes.",
                "resolution_required": "Author-reported fitted constants or a prospectively declared fit specification fixed before inspecting these curves.",
            },
            {
                "gate": "exact_sequence_label_mapping",
                "affected_labels": ["K1", "K2", "K3", "K31"],
                "reason_code": "discovery_sequences_not_deposited_or_identified",
                "reason": "Neither the workbook nor training_seqs.fasta maps the four discovery labels to exact sequences or accessions.",
                "resolution_required": "Author-deposited K1/K2/K3/K31 amino-acid sequences or an unambiguous label-to-accession table.",
            },
        ],
        "curation_effect": {
            "candidate_records_emitted": 0,
            "candidate_labels_created": False,
            "mmseqs_scope": "Pf-K only; the four unmapped labels cannot be searched without guessing",
            "registry_status": "scientifically_blocked",
        },
        "model_predictions_run": False,
    }

    homology_audit = SOURCE / "homology-audit.json"
    hit_path = SOURCE / "homology" / "homology_hits.tsv"
    if hit_path.exists():
        hit_lines = [line for line in hit_path.read_text(encoding="utf-8").splitlines() if line]
        homology = {
            "audited_on": "2026-07-27",
            "scope": "Pf-K only; K1, K2, K3, and K31 lack exact sequence mappings",
            "method": "MMseqs2 easy-search",
            "mmseqs_version": "5d152c612b6ad2a56f657b7a02c127eceaea2a75",
            "command": "mmseqs easy-search pf-k.fasta unikp_reference.fasta homology_hits.tsv TMP --min-seq-id 0.3 -c 0.8 --cov-mode 0 --format-output query,target,fident,alnlen,qcov,tcov,evalue,bits",
            "min_identity": 0.3,
            "coverage": 0.8,
            "coverage_mode": 0,
            "development_target_sha256": "69cbc91b88b69f7d4e8fe97f55d7ab157d7c438795989cd6a18fdb0ce28ab2f9",
            "query_sha256": sha256(SOURCE / "homology" / "pf-k.fasta"),
            "homology_hits_sha256": sha256(hit_path),
            "hit_count": len(hit_lines),
            "hit_queries": sorted({line.split("\t")[0] for line in hit_lines}),
            "hit_targets": sorted({line.split("\t")[1] for line in hit_lines}),
            "pf_k_excluded_homology": bool(hit_lines),
            "unmapped_labels_not_searched": ["K1", "K2", "K3", "K31"],
            "accepted_records_after_homology": 0,
            "model_predictions_run": False,
        }
        homology_audit.write_text(json.dumps(homology, indent=2) + "\n", encoding="ascii")
        output["homology"] = homology
    provenance = {
        "schema_version": 1,
        "source_id": "zenodo-10517668",
        "dataset_doi": "10.5281/zenodo.10517668",
        "stable_record_url": "https://zenodo.org/records/10517668",
        "article_doi": "10.1093/nar/gkae1245",
        "article_published": "2025-01-07",
        "dataset_license": "CC-BY-4.0",
        "kinetics_source": "raw/Fig2_raw_data.xlsx, sheet Fig. 2b, replicate cells B22:O50",
        "sequence_source": "raw/training_seqs.fasta",
        "source_sha256": observed_hashes,
        "auditor": "scripts/audit_zenodo_10517668.py",
        "normalization": "None; diagnostic fits are retained only in blocker.json and are not labels",
        "candidate_records_emitted": 0,
        "claim_boundary": "Scientifically blocked source; not part of the temporal curation pool",
        "model_predictions_run": False,
    }
    (SOURCE / "provenance.json").write_text(
        json.dumps(provenance, indent=2) + "\n", encoding="ascii"
    )
    (SOURCE / "blocker.json").write_text(json.dumps(output, indent=2) + "\n", encoding="ascii")


if __name__ == "__main__":
    main()
